from Pakiet import funkcje as functions
from Inicjator_polaczenia import acc, act
import openpyxl


''' STREFA KONFIGURACJI '''
excel_file_path = r'Inne\Arkusze\Dane_obiektow.xlsx'
sheet_name = "Sheet"
aliases = {}

# Załadowanie arkusza
wb = openpyxl.load_workbook(excel_file_path)
sheet = wb[sheet_name]

# Pobranie identyfikatorów elementów i właściwości
headers = functions.unpack_values(sheet.iter_rows(min_row=2, max_row=2, min_col=2))
properties_names = list(map(lambda x: aliases[x], headers)) if aliases else headers
properties_ids = functions.get_properties_ids(properties_names)
guids = functions.wrap_values(sheet.iter_cols(max_col=1, min_row=3))

# pobranie wartości właściwości w projekcie
properties = acc.GetPropertyValuesOfElements(guids, properties_ids)

wrapped_values = []

''' Przetwarzanie danych '''
# uzupełnienie kolekcji wrapped_values nowymi wartościami właściwości wskazanymi w arkuszu
for i, row in enumerate(sheet.iter_rows(min_row=3, min_col=2)):
    values = [cell.value for cell in row]
    for j, cell in enumerate(row):
        value = cell.value
        guid = guids[i]
        property_value = properties[i].propertyValues[j].propertyValue
        property_value.status = "normal"
        property_value.value = value
        wrapped_value = act.ElementPropertyValue(guids[i], properties_ids[j], property_value)
        wrapped_values.append(wrapped_value)

''' Monitorowanie wykonania poleceń'''
# zebranie i wyświetlenie w konsoli komunikatów o błędach
logs = acc.SetPropertyValuesOfElements(wrapped_values)
errors = [f"Error code {log.error.code}: {log.error.message}" for log in logs if not log.success]
if errors:
    print('The following problems have occurred:')
    for error in errors:
        print(error)
else:
    print('All operations successfully completed')
