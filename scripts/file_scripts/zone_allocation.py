from data_tools.connection_init import acc, act, acu
from data_tools import functions
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
import os


''' KONFIGURACJA SKRYPTU '''
fileIFC = "/Volumes/05_Modelowanie/Python skrypty/Inne/Pliki IFC/4 modelowanie.ifc"  # Ścieżka pliku IFC
outputFileName = "Zestawienie ścian według pomieszczeń.xlsx"  # Nazwa wyjściowego arkusza Excel
outputFolder = "/Volumes/05_Modelowanie/Python skrypty/Inne/Arkusze"  # Ścieżka wyjściowego arkusza
typesOfElements = ['Ściana', 'Przegroda']  # rozpatrywane typy elementów
StatesOfCurtainWall = (True, False)

''' Sprawdzanie dostępności właściwości '''
allProperties = acc.GetAllPropertyNames()
neededProperties = (
    ['Do skryptów', 'Nazwa'], ['Do skryptów', 'Kondygnacja'], ['Do skryptów', 'Materiał'],
    ['Do skryptów', 'Przydzielone'])
run = functions.is_all_properties_available(allProperties, neededProperties)
if run:

    ''' Ściąganie informacji z pliku IFC '''
    plik = open(fileIFC)
    text = plik.read()
    plik.close()
    x = text.find("#127=")
    text = text[x:]
    setOfAllLines = text.split("\n")
    spaces = [linia for linia in setOfAllLines if linia.find('IFCSPACE(') != -1]
    zonesInfo = [functions.generate_zoneInfo(space) for space in spaces]
    splittedText = text.split("IFCSPACE(")
    setOfLines = [text.split("\n") for text in splittedText]
    setOflines = setOfLines[:-1]
    wykaz = {}

    i = 0
    while i < len(setOflines):
        info = setOflines[i]
        isInfoAboutStory = False
        index = None
        j = 0
        for line in info:
            if line.find("IFCBUILDINGSTOREY") != -1:
                isInfoAboutStory = True
                index = j
            j += 1
        if isInfoAboutStory:
            info = info[index + 1:]
        rawDirections = [linia for linia in info if linia.find("IFCDIRECTION") != -1]
        rawDirectionInfo = rawDirections[-2]
        directionInfo = functions.check_direction(rawDirectionInfo)
        isPositiveX = directionInfo[0]
        isPositiveY = directionInfo[1]
        isReversed = directionInfo[2]

        rawPoints = [linia for linia in info if linia.find("IFCCARTESIANPOINT") != -1]
        points = [functions.przerabiarka(rawPoint) for rawPoint in rawPoints]
        startPoint = points[0][:-1]
        neededPoints = [point for point in points if len(point) == 2]
        polygonZone = functions.generate_polygonZone(startPoint, neededPoints, isPositiveX, isPositiveY, isReversed)
        wykaz[zonesInfo[i]] = polygonZone
        i += 1

    ''' Ściąganie informacji z projektu Archicada '''
    typesElements = functions.translate_types(typesOfElements)
    elements = []
    for type in typesElements:
        set = acc.GetElementsByType(type)
        elements += set
    zones = acc.GetElementsByType("Zone")
    propertyId1 = acu.GetBuiltInPropertyId('Zone_ZoneName')
    propertyId2 = acu.GetBuiltInPropertyId('Zone_ZoneNumber')
    propertyId3 = acu.GetBuiltInPropertyId('General_Type')
    propertyId4 = acu.GetBuiltInPropertyId('General_ElementID')
    propertyId5 = acu.GetUserDefinedPropertyId("Do skryptów", "Kondygnacja")
    propertyId6 = acu.GetUserDefinedPropertyId("Do skryptów", "Materiał")
    propertyId7 = acu.GetBuiltInPropertyId('General_SlantAngle')
    propertyId8 = acu.GetBuiltInPropertyId('General_Thickness')
    propertyId9 = acu.GetUserDefinedPropertyId('Do skryptów', 'Przydzielone')
    boundingBoxes = acc.Get2DBoundingBoxes(zones)
    boundingBoxes2 = acc.Get2DBoundingBoxes(elements)
    zoneNames = acc.GetPropertyValuesOfElements(zones, [propertyId1])
    zoneNumbers = acc.GetPropertyValuesOfElements(zones, [propertyId2])
    types = acc.GetPropertyValuesOfElements(elements, [propertyId3])
    Ids = acc.GetPropertyValuesOfElements(elements, [propertyId4])
    storyOfZone = acc.GetPropertyValuesOfElements(zones, [propertyId5])
    storyNumbers = acc.GetPropertyValuesOfElements(elements, [propertyId5])
    composites = acc.GetPropertyValuesOfElements(elements, [propertyId6])
    slantAngles = acc.GetPropertyValuesOfElements(elements, [propertyId7])
    thicknesses = acc.GetPropertyValuesOfElements(elements, [propertyId8])
    i = j = 0
    printedTexts = []
    printedPolygons = []
    assignedElements = []

    for box in boundingBoxes:
        zoneName = zoneNames[i].propertyValues[0].propertyValue.value
        zoneNumber = zoneNumbers[i].propertyValues[0].propertyValue.value
        prefix = f'{zoneNumber} {zoneName}'
        storyNumber = storyOfZone[i].propertyValues[0].propertyValue.value
        if prefix in wykaz:
            polygonZone = wykaz[prefix]
        else:
            i += 1
            continue
        j = 0
        for elembox in boundingBoxes2:
            storyNumber2 = storyNumbers[j].propertyValues[0].propertyValue.value
            typ = types[j].propertyValues[0].propertyValue.value
            xMin2 = elembox.boundingBox2D.xMin * 100
            xMax2 = elembox.boundingBox2D.xMax * 100
            yMin2 = elembox.boundingBox2D.yMin * 100
            yMax2 = elembox.boundingBox2D.yMax * 100
            if typ == "Przegroda strukturalna":
                slantAngle = slantAngles[j].propertyValues[0].propertyValue.value
                if slantAngle != 0:  # jeśli przegroda jest "stojąca"
                    if StatesOfCurtainWall[0] is True:  # jeśli ustawiono uwzględnianie przegród stojących
                        thickness = thicknesses[j].propertyValues[0].propertyValue.value * 100
                        lengthx = xMax2 - xMin2
                        lengthy = yMax2 - yMin2
                        lengths = [lengthx, lengthy]
                        lengths.sort()
                        if lengths[0] < 1:
                            if lengths[0] == lengthx:
                                xMin2 -= thickness / 2
                                xMax2 += thickness / 2
                            else:
                                yMin2 -= thickness / 2
                                yMax2 += thickness / 2
                    else:
                        j += 1
                        continue
                else:
                    if StatesOfCurtainWall[1] is False:  # jeśli ustawiono nieuwzględnianie przegród leżących
                        j += 1
                        continue
            vertices2 = [xMin2, xMax2, yMin2, yMax2]
            polygonElement = functions.generate_polygon_element(vertices2)
            state = functions.multi_is_inside(polygonZone, polygonElement)
            if state is True:  # jeśli element znajduje się w sprawdzanej strefie w 2D
                if storyNumber == storyNumber2:  # jeśli element znajduje się na tej samej kondygnacji co strefa
                    guid = elements[j]
                    if guid not in assignedElements:
                        assignedElements.append(guid)
                    Id = Ids[j].propertyValues[0].propertyValue.value
                    material = composites[j].propertyValues[0].propertyValue.value
                    text = f'Nazwa i numer strefy: {zoneNumber} {zoneName}'
                    if text not in printedTexts:
                        printedTexts.append(text)
                    text2 = f"  {typ}; {Id}; {material}"
                    printedTexts.append(text2)
            j += 1
        i += 1

    ''' Ustawienia kolejności stref '''
    relations = {}
    values = []
    isFirst = True
    key = None
    for text in printedTexts:
        if text[0] != " ":
            if isFirst:
                key = text[22:]
                isFirst = False
            else:
                relations[key] = values
                values = []
                key = text[22:]
        else:
            values.append(text[2:])
    if key != None:
        relations[key] = values

    relations = dict(sorted(relations.items()))
    for relation in relations:
        print("Nazwa i numer strefy:", relation)
        elem = relations[relation]
        for el in elem:
            print(" ", el)

    ''' Zapisywanie informacji do pliku Excel '''
    scriptFolder = os.path.dirname(os.path.realpath(__file__))
    wb = Workbook()
    sheet = wb.active
    font = Font(name='Arial', size=14)
    titlefont = Font(name='Arial', size=17)
    yes = Side(border_style='thin', color='00000000')
    border1 = Border(left=yes, right=yes, top=yes, bottom=yes)
    border2 = Border(left=yes, right=yes)
    border3 = Border(left=yes, right=yes, top=None, bottom=yes)
    settings = functions.set_field_parameters

    title1 = 'Numer i nazwa strefy'
    title2 = 'Typ elementu'
    title3 = 'ID'
    title4 = 'Materiał'
    settings(sheet, 2, 2, title1, titlefont, border1)
    settings(sheet, 2, 3, title2, titlefont, border1)
    settings(sheet, 2, 4, title3, titlefont, border1)
    settings(sheet, 2, 5, title4, titlefont, border1)

    cellsToMerge = []
    zonesInfo.sort()
    row = 3
    for zone in zonesInfo:
        if zone in relations:
            Elements = relations[zone]
        else:
            continue
        if len(Elements) > 0:
            startRow = row
            printedSplits = []
            for Element in Elements:
                split = Element.split(";")
                type, id, compositeName = split
                if split not in printedSplits:
                    settings(sheet, row, 3, type, font, border2)
                    settings(sheet, row, 4, id, font, border2)
                    settings(sheet, row, 5, compositeName, font, border2)
                    row += 1
                    printedSplits.append(split)
            endRow = row - 1
            settings(sheet, startRow, 2, zone, font, border1)
            sheet.merge_cells(start_row=startRow, end_row=endRow, start_column=2, end_column=2)
            sheet.cell(endRow, 3).border = border3
            sheet.cell(endRow, 4).border = border3
            sheet.cell(endRow, 5).border = border3
        else:
            sheet.cell(row, 2).value = zone
            row += 1
    functions.auto_fit_worksheet_columns(sheet)
    excelFilePath = os.path.join(outputFolder, outputFileName)
    try:
        wb.save(excelFilePath)
    except BlockingIOError:
        print(
            "\nMasz otwarty plik Excel, który uniemożliwia zapisanie arkusza. \nZamknij go i uruchom skrypt ponownie.")
    else:
        acu.OpenFile(excelFilePath)
        if os.path.exists(excelFilePath):
            print("Zapisano plik Excel.")

        ''' Oznaczenie nieprzydzielonych elementów '''
        aloneElements = [element for element in elements if element not in assignedElements]
        i = 0
        if aloneElements:
            print()
            print("Niektórych elementów nie przydzielono do zadnej strefy.")
            elementPropertyValues = []
            propertyValue = act.NormalStringPropertyValue('NIE', 'string', 'normal')
            for element in aloneElements:
                elementId = element.elementId
                x = act.ElementPropertyValue(elementId, propertyId9, propertyValue)
                elementPropertyValues.append(x)
                i += 1
            i = 0
            propertyValue = act.NormalStringPropertyValue('TAK', 'string', 'normal')
            for element in assignedElements:
                elementId = element.elementId
                x = act.ElementPropertyValue(elementId, propertyId9, propertyValue)
                elementPropertyValues.append(x)
                i += 1

        y = acc.SetPropertyValuesOfElements(elementPropertyValues)
        if y[0].success:
            print("Przydzielono im wartość 'NIE' dla właściwości 'Przydzielone'.")
        else:
            print(
                f"Operacja przydzielania wartości do właściwości \n 'Przydzielone' zakończona niepowodzeniem.\n{y[0].error}")
