from data_tools.connection_init import acc, acu
from data_tools.functions import auto_fit_worksheet_columns, set_field_parameters
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment, numbers
import os


''' Konfiguracja skryptu '''
output_file_name = r"Zestawienie mebli według pomieszczeń2.xlsx"  # Nazwa wyjściowego arkusza Excel
output_folder = r"C:\Users\asus\Downloads\Python skrypty\Inne\Arkusze"  # Ścieżka wyjściowego arkusza
properties = ['IdAndCategories_RelatedZoneNumber', 'IdAndCategories_RelatedZoneName', 'General_ElementID',
              ('Specyfikacja produktu', 'Nazwa'), 'Geometry_ObjectLength', 'General_Width',
                'General_Height',  ('Specyfikacja produktu', 'Cena zakupu'),]

''' Pobieranie wartości właściwości'''
objects = acc.GetElementsByType('Object') # elementy sprawdzane w skrypcie
all_properties_ids = [acu.GetBuiltInPropertyId(property) if type(property) is str else acu.GetUserDefinedPropertyId(*property)
                      for property in properties]
all_properties_values = acc.GetPropertyValuesOfElements(objects, all_properties_ids)

''' Przydzielanie obiektów do pomieszczeń'''
room_layout = {}
for property in all_properties_values:
    property_set = [property_data.propertyValue.value if property_data.propertyValue.status == 'normal' else '-' 
                    for property_data in property.propertyValues]
    zone = ' '.join(property_set[:2])
    if zone.isspace():  # ignoruje obiekty nieprzydzielone do żadnej strefy
        continue
    x, y, z = property_set[4:7]
    dimensions = f"{(100*x):.0f} x {(100*y):.0f} x {(100*z):.0f}"
    properties = tuple(property_set[2:4] + [dimensions] + [property_set[7]])
    room_objects = room_layout.get(zone, [])
    room_objects.append(properties)
    room_layout[zone] = room_objects
zones = sorted(room_layout)

''' Zapisywanie informacji do pliku Excel '''

# definicja arkusza oraz powiązanych z nim ustawień
scriptFolder = os.path.dirname(os.path.realpath(__file__))
wb = Workbook()
sheet = wb.active
font = Font(name='Arial', size=14)
title_font = Font(name='Arial', size=16)
title_align = Alignment('center')
thin_border = Side(border_style='thin', color='00000000')
border1 = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
border2 = Border(left=thin_border, right=thin_border)
border3 = Border(left=thin_border, right=thin_border, bottom=thin_border)
border4 = Border(right=thin_border, top=thin_border, bottom=thin_border)
titles = ['Numer i nazwa strefy', 'ID', 'Nazwa', 'Wymiary [cm]', 'Ilość', 'Cena [zł]', 'Wartość [zł]']

# ustawienie nagłówków kolumn
for i, title in enumerate(titles):
    set_field_parameters(2, i+2, title, title_font, border1, alignment=title_align)

# wstawianie wartości w komórki związane z kolejnymi strefami
summary_indexes = [] # numery wierszy zawierających sumy wartości obiektów w poszczególnych strefach
row = 3
for zone in zones:
    elements = room_layout[zone]
    if elements:
        elem_counter = dict(Counter(elements))
        startRow = row
        for element, amount in elem_counter.items():
            element = element[:3] + (amount,) + element[3:] + (f'=F{row}*G{row}',)
            for i, property in enumerate(element):
                number_format = None if i < 4 else numbers.FORMAT_NUMBER_00
                set_field_parameters(row, i+3, property, font, border2, number_format)
            row += 1
        row += 1
        endRow = row - 1
        set_field_parameters(startRow, 2, zone, font, border1)
        sheet.merge_cells(start_row=startRow, end_row=endRow-1, start_column=2, end_column=2)
        
        # przygotowanie ostatniego rzędu:
        for i in range(2, len(titles)+2):
            border = border4 if i == 3 else border1
            fg_color = 'FFDA9694' if i < len(titles) else 'FFC4D79B'
            fill = PatternFill(fgColor=fg_color, fill_type='solid')
            set_field_parameters(endRow, i, border=border, fill=fill)
        set_field_parameters(endRow, 7, "Suma:", font)
        formula = f'=SUM(H{startRow}:H{endRow-1})'
        set_field_parameters(endRow, 8, formula, font, number_format=numbers.FORMAT_NUMBER_00)
        sheet.merge_cells(start_row=endRow, end_row=endRow, start_column=2, end_column=6)
        summary_indexes.append(endRow)
    else:
        set_field_parameters(row, 2, zone)
        row += 1

# Pełne podsumowanie pod główną tabelą
fill = PatternFill(fgColor='FF76933C', fill_type='solid')
set_field_parameters(row, 7, "Pełna suma:", font=font, border=border1, fill=fill)
formula = '=' + '+'.join([f'H{index}' for index in summary_indexes])
set_field_parameters(row, 8, formula, font, border1, numbers.FORMAT_NUMBER_00, fill)

auto_fit_worksheet_columns(sheet)
excelFilePath = os.path.join(output_folder, output_file_name)

try:
    wb.save(excelFilePath)
except PermissionError:
    print(
        "\nMasz otwarty plik Excel, który uniemożliwia zapisanie arkusza. \nZamknij go i uruchom skrypt ponownie.")
else:
    acu.OpenFile(excelFilePath)
    if os.path.exists(excelFilePath):
        print("Zapisano plik Excel.")
