from Pakiet import funkcje as functions
from openpyxl import Workbook
from Inicjator_polaczenia import acc, acu
import os


''' KONFIGURACJA SKRYPTU '''
excelFilePath = r'Inne\Arkusze\Dane_scian.xlsx'
properties_names = ['General_ElementID','Category_Position', 'General_3DLength', 'General_Width', 'General_Height']

elements = acc.GetSelectedElements()
headers = ['Element_Guid'] + properties_names
all_properties_ids = functions.get_properties_ids(properties_names)
property_objects = acc.GetPropertyValuesOfElements(elements, all_properties_ids)

wb = Workbook()
sheet = wb.active

for i, header in enumerate(headers, start=1):
    sheet.cell(2, i).value = header

for i, element in enumerate(elements, start=3):
    guid = str(element.elementId.guid)
    sheet.cell(i,1).value = guid


for i, property_object in enumerate(property_objects, start=3):
    property_values = property_object.propertyValues
    for j, property_value in enumerate(property_values, start=2):
        status = property_value.propertyValue.status
        if status == 'normal':
            value = property_value.propertyValue.value
            if hasattr(value, 'nonLocalizedValue'):
                value = value.nonLocalizedValue
            sheet.cell(i, j).value = value

functions.auto_fit_worksheet_columns(sheet)

try:
    wb.save(excelFilePath)
except PermissionError:
    print(
        "\nMasz otwarty plik Excel, który uniemożliwia zapisanie arkusza. \nZamknij go i uruchom skrypt ponownie.")
else:
    acu.OpenFile(excelFilePath)
    if os.path.exists(excelFilePath):
        print("Zapisano plik Excel.")
